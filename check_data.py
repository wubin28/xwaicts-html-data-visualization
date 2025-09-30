import pandas as pd

# 尝试不同的方式读取Excel文件
# 方式1: 跳过第一行，使用第二行作为列名
df1 = pd.read_excel('first-80-rows-agentic_ai_performance_dataset_20250622.xlsx', skiprows=1)
print('=== 方式1: 跳过第一行 ===')
print('数据形状:', df1.shape)
print('\n列名:', df1.columns.tolist())
print('\n前5行数据:')
print(df1.head())

# 方式2: 不使用任何行作为列名
df2 = pd.read_excel('first-80-rows-agentic_ai_performance_dataset_20250622.xlsx', header=None)
print('\n=== 方式2: 不使用任何行作为列名 ===')
print('数据形状:', df2.shape)
print('\n前5行数据:')
print(df2.head())

# 方式3: 尝试使用openpyxl直接读取
from openpyxl import load_workbook

print('\n=== 方式3: 使用openpyxl直接读取 ===')
wb = load_workbook('first-80-rows-agentic_ai_performance_dataset_20250622.xlsx')
sheet = wb.active
print('工作表名称:', sheet.title)
print('最大行数:', sheet.max_row)
print('最大列数:', sheet.max_column)

# 打印前5行数据
print('\n前5行数据:')
for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
    print(row)